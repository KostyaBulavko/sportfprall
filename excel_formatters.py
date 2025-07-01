# excel_formatters.py
"""
Форматирование и стили для Excel файлов
"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from excel_config import get_ordered_headers, get_column_widths


def apply_cell_formatting(ws, max_row, max_col):
    """Застосування рамок і вирівнювання до всіх комірок"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # Не заголовки
                cell.alignment = Alignment(horizontal="left", vertical="center")


def format_headers(ws, headers):
    """Форматирование заголовков"""
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_column_widths(ws, fields_list):
    """Применение ширины колонок"""
    column_widths = get_column_widths(fields_list)
    headers = get_ordered_headers(fields_list)

    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        width = column_widths.get(header, 15)  # По умолчанию 15
        ws.column_dimensions[column_letter].width = width