# koshtorys.py
# -*- coding: utf-8 -*-

import os
import tkinter.filedialog as filedialog
import tkinter.messagebox as messagebox
import sys
import traceback
from datetime import datetime
import shutil

from openpyxl.styles import Font
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string


from copy import copy


# Спробуємо імпортувати openpyxl для роботи з Excel
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import RowDimension

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Увага: openpyxl не встановлено. Буде створено текстовий файл замість Excel.")

# Импортируем необходимые функции
try:
    from error_handler import log_error_koshtorys as log_error
    from text_utils import number_to_ukrainian_text
except ImportError:
    print("Критическая ошибка: не удалось импортировать модули error_handler или text_utils для koshtorys.py")


    def log_error(exc_type, exc_value, exc_traceback, error_log="error.txt"):
        print(f"Ошибка (koshtorys log): {exc_value}")
        traceback.print_exception(exc_type, exc_value, exc_traceback)
        messagebox.showerror("Ошибка", f"Произошла ошибка (koshtorys): {str(exc_value)}")


    def number_to_ukrainian_text(amount):
        return f"{amount} грн (заглушка)"


def get_entry_value(entries_dict, field_name, default=""):
    """Отримує значення з Entry віджета"""
    if field_name in entries_dict:
        widget = entries_dict[field_name]
        if hasattr(widget, 'get'):
            return widget.get().strip()
    return default


def convert_to_number(value_str):
    """Конвертує строку в число"""
    if not value_str:
        return 0.0

    try:
        # Замінюємо кому на крапку та видаляємо пробіли
        clean_str = str(value_str).replace(",", ".").replace(" ", "")
        return float(clean_str)
    except ValueError:
        return 0.0


def get_товар_name_from_entries(entries):
    """Знаходить назву товару серед різних можливих полів"""
    # Список можливих полів для назви товару (в порядку пріоритету)
    товар_fields = ["товар", "назва", "найменування", "предмет", "послуга", "робота", "матеріал", "дк"]

    for field in товар_fields:
        value = get_entry_value(entries, field)
        if value:  # Якщо поле не пусте
            return value

    return ""


def get_товари_from_blocks(document_blocks):
    """Збирає всі товари з блоків документів"""
    товари = []

    for block in document_blocks:
        entries = block.get("entries", {})

        # Шукаємо назву товару серед різних можливих полів
        товар = get_товар_name_from_entries(entries)
        кількість = get_entry_value(entries, "кількість")

        # Шукаємо ціну серед можливих варіантів
        ціна = (get_entry_value(entries, "ціна за одиницю") or
                get_entry_value(entries, "ціна") or
                get_entry_value(entries, "вартість за одиницю") or
                get_entry_value(entries, "вартість одиниці"))

        # Шукаємо одиницю виміру
        одиниця = (get_entry_value(entries, "одиниця виміру") or
                   get_entry_value(entries, "од") or
                   get_entry_value(entries, "шт") or
                   "шт")  # за замовчуванням

        if товар:  # Якщо є назва товару
            товари.append({
                "товар": товар,
                "кількість": кількість,
                "ціна за одиницю": ціна,
                "одиниця виміру": одиниця
            })

    return товари


def is_merged_cell(sheet, row, col):
    """Перевіряє чи є ячейка об'єднаною"""
    try:
        from openpyxl.worksheet.merge import MergedCell
        cell = sheet.cell(row, col)
        return isinstance(cell, MergedCell)
    except:
        return False


def get_merged_cell_top_left(sheet, row, col):
    """Отримує координати верхньої лівої ячейки об'єднаної групи"""
    try:
        for merged_range in sheet.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                return merged_range.min_row, merged_range.min_col
        return row, col
    except:
        return row, col


def safe_set_cell_value(sheet, row, col, value):
    """Безпечно встановлює значення ячейки (враховує об'єднані ячейки)"""
    try:
        if is_merged_cell(sheet, row, col):
            # Для об'єднаної ячейки знаходимо верхню ліву ячейку
            top_row, left_col = get_merged_cell_top_left(sheet, row, col)
            target_cell = sheet.cell(top_row, left_col)
            #print(f"[DEBUG] Об'єднана ячейка ({row},{col}) -> встановлюємо значення в ({top_row},{left_col})")
        else:
            target_cell = sheet.cell(row, col)

        target_cell.value = value
        return True
    except Exception as e:
        print(f"[WARNING] Не вдалося встановити значення в ячейку ({row},{col}): {e}")
        return False


def copy_row_with_full_formatting(sheet, source_row_num, target_row_num):
    """Копіює повне форматування рядка включаючи висоту, стилі всіх ячейок"""
    try:
        max_col = sheet.max_column if sheet.max_column > 0 else 26

        # Копіюємо висоту рядка
        if source_row_num in sheet.row_dimensions:
            source_row_dim = sheet.row_dimensions[source_row_num]
            target_row_dim = sheet.row_dimensions[target_row_num]
            target_row_dim.height = source_row_dim.height
            target_row_dim.hidden = source_row_dim.hidden
            target_row_dim.outlineLevel = source_row_dim.outlineLevel
            target_row_dim.collapsed = source_row_dim.collapsed

        # Копіюємо всі ячейки з повним форматуванням
        for col_num in range(1, max_col + 1):
            source_cell = sheet.cell(source_row_num, col_num)
            target_cell = sheet.cell(target_row_num, col_num)

            # Перевіряємо чи цільова ячейка не є частиною об'єднання
            # Якщо так, пропускаємо її (об'єднання будуть скопійовані окремо)
            if is_merged_cell(sheet, target_row_num, col_num):
                continue

            # Копіюємо значення (якщо воно не формула і не пусте)
            if source_cell.value is not None and not str(source_cell.value).startswith('='):
                target_cell.value = source_cell.value

            # Копіюємо повний стиль через _style (найнадійніший спосіб)
            if hasattr(source_cell, '_style') and source_cell._style:
                target_cell._style = copy(source_cell._style)

            # Додатково копіюємо кожен стиль окремо для гарантії
            try:
                # Шрифт
                if source_cell.font:
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic,
                        vertAlign=source_cell.font.vertAlign,
                        underline=source_cell.font.underline,
                        strike=source_cell.font.strike,
                        color=copy(source_cell.font.color) if source_cell.font.color else None
                    )

                # Заливка
                if source_cell.fill and source_cell.fill.fill_type:
                    target_cell.fill = PatternFill(
                        fill_type=source_cell.fill.fill_type,
                        start_color=copy(source_cell.fill.start_color) if source_cell.fill.start_color else None,
                        end_color=copy(source_cell.fill.end_color) if source_cell.fill.end_color else None
                    )

                # Границі
                if source_cell.border:
                    target_cell.border = Border(
                        left=copy(source_cell.border.left) if source_cell.border.left else None,
                        right=copy(source_cell.border.right) if source_cell.border.right else None,
                        top=copy(source_cell.border.top) if source_cell.border.top else None,
                        bottom=copy(source_cell.border.bottom) if source_cell.border.bottom else None,
                        diagonal=copy(source_cell.border.diagonal) if source_cell.border.diagonal else None,
                        diagonal_direction=source_cell.border.diagonal_direction,
                        outline=copy(source_cell.border.outline) if source_cell.border.outline else None,
                        vertical=copy(source_cell.border.vertical) if source_cell.border.vertical else None,
                        horizontal=copy(source_cell.border.horizontal) if source_cell.border.horizontal else None
                    )

                # Вирівнювання
                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical,
                        text_rotation=source_cell.alignment.text_rotation,
                        wrap_text=source_cell.alignment.wrap_text,
                        shrink_to_fit=source_cell.alignment.shrink_to_fit,
                        indent=source_cell.alignment.indent
                    )

                # Формат чисел
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format

                # Захист
                if source_cell.protection:
                    target_cell.protection = copy(source_cell.protection)

            except Exception as style_error:
                print(f"[WARNING] Помилка копіювання стилю ячейки {col_num}: {style_error}")

        # Копіюємо об'єднані ячейки в рядку
        merged_ranges_to_copy = []
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row == source_row_num and merged_range.max_row == source_row_num:
                # Це об'єднання тільки в межах одного рядка
                merged_ranges_to_copy.append(merged_range)

        for merged_range in merged_ranges_to_copy:
            try:
                new_range = f"{get_column_letter(merged_range.min_col)}{target_row_num}:{get_column_letter(merged_range.max_col)}{target_row_num}"
                sheet.merge_cells(new_range)
                #print(f"[DEBUG] Скопійовано об'єднання в рядку: {new_range}")

                # Копіюємо значення в об'єднану ячейку
                source_merged_cell = sheet.cell(source_row_num, merged_range.min_col)
                target_merged_cell = sheet.cell(target_row_num, merged_range.min_col)
                if source_merged_cell.value is not None:
                    target_merged_cell.value = source_merged_cell.value

            except Exception as merge_error:
                print(f"[WARNING] Помилка копіювання об'єднання: {merge_error}")

        #print(f"[DEBUG] Скопійовано повне форматування з рядка {source_row_num} в рядок {target_row_num}")

    except Exception as e:
        print(f"[ERROR] Помилка копіювання форматування рядка: {e}")
        traceback.print_exc()


def insert_rows_for_products(sheet, товари_count):
    """Вставляє рядки для товарів з повним збереженням форматування рядків 33-40"""
    try:
        # Базовий рядок для товарів в шаблоні (рядок 32)
        base_product_row = 32
        # Початок критичної області з підсумками та підписами
        critical_start_row = 33
        # Кінець критичної області (рядки 33-40)
        critical_end_row = 40

        if товари_count <= 1:
            # Якщо товар один або менше, нічого не змінюємо
            #print(f"[DEBUG] Товарів {товари_count}, додаткові рядки не потрібні")
            return base_product_row

        # Кількість додаткових рядків, які потрібно вставити
        additional_rows = товари_count - 1

        #print(f"[DEBUG] Потрібно вставити {additional_rows} додаткових рядків для {товари_count} товарів")
        #print(f"[DEBUG] Базовий рядок товару: {base_product_row}")
        #print(f"[DEBUG] Критична область: {critical_start_row}-{critical_end_row}")

        max_col = sheet.max_column if sheet.max_column > 0 else 26

        # КРОК 1: Повне збереження критичної області (33-40) з усіма стилями та об'єднаннями
        #print(f"[DEBUG] Зберігаємо критичну область {critical_start_row}-{critical_end_row}")
        saved_critical_area = {}

        # Зберігаємо висоту рядків
        saved_row_dimensions = {}
        for row in range(critical_start_row, critical_end_row + 1):
            if row in sheet.row_dimensions:
                saved_row_dimensions[row] = {
                    'height': sheet.row_dimensions[row].height,
                    'hidden': sheet.row_dimensions[row].hidden,
                    'outlineLevel': sheet.row_dimensions[row].outlineLevel,
                    'collapsed': sheet.row_dimensions[row].collapsed,
                }

        # Зберігаємо об'єднані ячейки в критичній області
        saved_merged_cells = []
        for merged_range in list(sheet.merged_cells.ranges):
            if (critical_start_row <= merged_range.min_row <= critical_end_row or
                    critical_start_row <= merged_range.max_row <= critical_end_row):
                saved_merged_cells.append({
                    'range': str(merged_range),
                    'min_row': merged_range.min_row,
                    'max_row': merged_range.max_row,
                    'min_col': merged_range.min_col,
                    'max_col': merged_range.max_col
                })
                #print(f"[DEBUG] Збережено об'єднання: {merged_range}")

        # Зберігаємо всі ячейки критичної області
        for row in range(critical_start_row, critical_end_row + 1):
            saved_critical_area[row] = {}
            for col in range(1, max_col + 1):
                cell = sheet.cell(row, col)

                # Зберігаємо повну інформацію про ячейку
                saved_critical_area[row][col] = {
                    'value': cell.value,
                    'style': copy(cell._style) if hasattr(cell, '_style') else None,
                    'font': copy(cell.font) if cell.font else None,
                    'fill': copy(cell.fill) if cell.fill else None,
                    'border': copy(cell.border) if cell.border else None,
                    'alignment': copy(cell.alignment) if cell.alignment else None,
                    'number_format': cell.number_format,
                    'protection': copy(cell.protection) if cell.protection else None,
                    'hyperlink': cell.hyperlink,
                    'comment': cell.comment
                }

        #print(f"[DEBUG] Збережено {len(saved_critical_area)} рядків з повним форматуванням")

        # КРОК 2: Видаляємо об'єднання в критичній області перед вставкою
        merged_to_remove = []
        for merged_range in list(sheet.merged_cells.ranges):
            if (critical_start_row <= merged_range.min_row <= critical_end_row or
                    critical_start_row <= merged_range.max_row <= critical_end_row):
                merged_to_remove.append(merged_range)

        for merged_range in merged_to_remove:
            sheet.unmerge_cells(str(merged_range))
            #print(f"[DEBUG] Тимчасово видалено об'єднання: {merged_range}")

        # КРОК 3: Вставляємо додаткові рядки ПІСЛЯ базового рядка товару (32)
        # Це автоматично зсуне всі рядки нижче (включаючи 33-40)
        for i in range(additional_rows):
            insert_position = base_product_row + 1 + i
            sheet.insert_rows(insert_position, 1)
            #print(f"[DEBUG] Вставлено рядок товару на позиції {insert_position}")

            # Копіюємо форматування з базового рядка товару
            copy_row_with_full_formatting(sheet, base_product_row, insert_position)

        # КРОК 4: Розраховуємо нові позиції для критичної області
        new_critical_start = critical_start_row + additional_rows
        new_critical_end = critical_end_row + additional_rows

        #print(f"[DEBUG] Нові позиції критичної області: {new_critical_start}-{new_critical_end}")

        # КРОК 5: Відновлюємо критичну область на нових позиціях
        for original_row in range(critical_start_row, critical_end_row + 1):
            new_row = original_row + additional_rows

            # Відновлюємо висоту рядка
            if original_row in saved_row_dimensions:
                sheet.row_dimensions[new_row].height = saved_row_dimensions[original_row]['height']
                sheet.row_dimensions[new_row].hidden = saved_row_dimensions[original_row]['hidden']
                sheet.row_dimensions[new_row].outlineLevel = saved_row_dimensions[original_row]['outlineLevel']
                sheet.row_dimensions[new_row].collapsed = saved_row_dimensions[original_row]['collapsed']

            # Відновлюємо всі ячейки в рядку
            for col in range(1, max_col + 1):
                if original_row in saved_critical_area and col in saved_critical_area[original_row]:
                    cell_data = saved_critical_area[original_row][col]
                    target_cell = sheet.cell(new_row, col)

                    # Відновлюємо значення
                    target_cell.value = cell_data['value']

                    # Відновлюємо стилі
                    if cell_data['style']:
                        target_cell._style = cell_data['style']

                    if cell_data['font']:
                        target_cell.font = cell_data['font']
                    if cell_data['fill']:
                        target_cell.fill = cell_data['fill']
                    if cell_data['border']:
                        target_cell.border = cell_data['border']
                    if cell_data['alignment']:
                        target_cell.alignment = cell_data['alignment']
                    if cell_data['number_format']:
                        target_cell.number_format = cell_data['number_format']
                    if cell_data['protection']:
                        target_cell.protection = cell_data['protection']
                    if cell_data['hyperlink']:
                        target_cell.hyperlink = cell_data['hyperlink']
                    if cell_data['comment']:
                        target_cell.comment = cell_data['comment']

        #print(f"[DEBUG] Відновлено всі ячейки критичної області")

        # КРОК 6: Відновлюємо об'єднані ячейки на нових позиціях
        for merged_info in saved_merged_cells:
            try:
                # Розраховуємо нові координати об'єднання
                new_min_row = merged_info['min_row'] + additional_rows
                new_max_row = merged_info['max_row'] + additional_rows
                new_min_col = merged_info['min_col']
                new_max_col = merged_info['max_col']

                # Створюємо нове об'єднання
                new_range = f"{get_column_letter(new_min_col)}{new_min_row}:{get_column_letter(new_max_col)}{new_max_row}"
                sheet.merge_cells(new_range)
                #print(f"[DEBUG] Відновлено об'єднання: {merged_info['range']} -> {new_range}")

            except Exception as merge_error:
                print(f"[WARNING] Помилка відновлення об'єднання {merged_info['range']}: {merge_error}")

        # КРОК 7: Очищаємо старі рядки критичної області (якщо вони ще залишились)
        # Це може статися якщо була накладка в позиціях
        for row in range(critical_start_row, critical_start_row + additional_rows):
            for col in range(1, max_col + 1):
                try:
                    cell = sheet.cell(row, col)
                    # Очищуємо тільки якщо ячейка не є частиною нового товару
                    if row > base_product_row + товари_count:
                        cell.value = None
                except:
                    pass

        #print(f"[DEBUG] Успішно вставлено {additional_rows} рядків товарів")
        # print(
        #     f"[DEBUG] Критична область {critical_start_row}-{critical_end_row} збережена і переміщена на {new_critical_start}-{new_critical_end}")

        return base_product_row

    except Exception as e:
        print(f"[ERROR] Помилка при вставці рядків: {e}")
        traceback.print_exc()
        return base_product_row


def safe_write_to_cell(sheet, cell_reference, value):
    """Безпечно записує значення в ячейку, враховуючи об'єднані ячейки"""
    try:
        #print(f"[DEBUG] Спроба запису '{value}' в ячейку {cell_reference}")

        # Розбираємо референс ячейки (наприклад, "K39" -> row=39, col=11)
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        col_letter, row = coordinate_from_string(cell_reference)
        col_num = column_index_from_string(col_letter)

        #print(f"[DEBUG] Координати: рядок={row}, колонка={col_num} ({col_letter})")

        # Отримуємо ячейку
        cell = sheet.cell(row, col_num)
        #print(f"[DEBUG] Поточне значення ячейки: '{cell.value}'")

        # Перевіряємо чи є ячейка об'єднаною
        is_merged = is_merged_cell(sheet, row, col_num)
        #print(f"[DEBUG] Ячейка об'єднана: {is_merged}")

        if is_merged:
            # Для об'єднаної ячейки знаходимо верхню ліву ячейку
            top_row, left_col = get_merged_cell_top_left(sheet, row, col_num)
            target_cell = sheet.cell(top_row, left_col)
            #print(f"[DEBUG] Об'єднана ячейка ({row},{col_num}) -> встановлюємо значення в ({top_row},{left_col})")
            #print(f"[DEBUG] Поточне значення верхньої лівої ячейки: '{target_cell.value}'")
        else:
            target_cell = cell
            #print(f"[DEBUG] Звичайна ячейка, записуємо безпосередньо")

        # Встановлюємо значення
        old_value = target_cell.value
        target_cell.value = value
        new_value = target_cell.value

        #print(f"[DEBUG] Змінено значення з '{old_value}' на '{new_value}'")

        # Перевіряємо чи запис успішний
        if str(target_cell.value) == str(value):
            #print(f"[DEBUG] ✓ Успішно записано '{value}' в ячейку {cell_reference}")
            return True
        else:
            print(
                f"[WARNING] ✗ Значення не співпадає після запису: очікувалось '{value}', отримано '{target_cell.value}'")
            return False

    except Exception as e:
        print(f"[ERROR] Помилка запису в ячейку {cell_reference}: {e}")
        import traceback
        traceback.print_exc()
        return False


def save_koshtorys_to_excel(захід, адреса, дата, товари, загальна_сума, сума_прописом):
    """Зберігає кошторис у Excel файл на основі шаблону з правильним заповненням"""

    def debug_template_structure(sheet):
        """Диагностика структуры шаблона"""

        # Проверяем содержимое ячеек вокруг целевых
        areas_to_check = [
            ("D12 область", range(10, 15), range(3, 6)),  # D12 область
            ("E14-E15 область", range(13, 17), range(4, 7))  # E14-E15 область
        ]

        for area_name, rows, cols in areas_to_check:
            print(f"\n[DEBUG] {area_name}:")
            for row in rows:
                for col in cols:
                    try:
                        cell = sheet.cell(row, col)
                        col_letter = get_column_letter(col)
                        value = cell.value
                        if value:
                            print(f"[DEBUG] {col_letter}{row}: '{value}'")
                    except:
                        pass

    def improved_fill_main_data(sheet, захід, адреса, дата):
        """Покращена функція заповнення основних даних з детальною діагностикою"""
        try:
            success_count = 0
            total_attempts = 0

            # Заповнюємо захід в D12
            if захід and захід.strip():
                total_attempts += 1
                methods = [
                    ('safe_write_to_cell', lambda: safe_write_to_cell(sheet, 'D12', захід)),
                    ('direct_cell_access', lambda: setattr(sheet['D12'], 'value', захід)),
                    ('cell_coordinate', lambda: setattr(sheet.cell(12, 4), 'value', захід))
                ]

                for method_name, method_func in methods:
                    try:
                        result = method_func()
                        actual_value = sheet['D12'].value

                        if actual_value and str(actual_value).strip() == str(захід).strip():
                            success_count += 1
                            break
                        else:
                            print(f"[DEBUG] ✗ Неуспіх методом {method_name}")

                    except Exception as e:
                        print(f"[DEBUG] Помилка методом {method_name}: {e}")
            else:
                print(f"[DEBUG] Захід пустий або None, пропускаємо")

            # Заповнюємо адресу в E14
            if адреса and адреса.strip():
                total_attempts += 1
                methods = [
                    ('safe_write_to_cell', lambda: safe_write_to_cell(sheet, 'E14', адреса)),
                    ('direct_cell_access', lambda: setattr(sheet['E14'], 'value', адреса)),
                    ('cell_coordinate', lambda: setattr(sheet.cell(14, 5), 'value', адреса))
                ]

                for method_name, method_func in methods:
                    try:
                        result = method_func()
                        actual_value = sheet['E14'].value

                        if actual_value and str(actual_value).strip() == str(адреса).strip():
                            success_count += 1
                            break
                        else:
                            print(f"[DEBUG] ✗ Неуспіх методом {method_name}")

                    except Exception as e:
                        print(f"[DEBUG] Помилка методом {method_name}: {e}")
            else:
                print(f"[DEBUG] Адреса пуста або None, пропускаємо")

            # Заповнюємо дату в E15
            if дата and дата.strip():
                total_attempts += 1
                methods = [
                    ('safe_write_to_cell', lambda: safe_write_to_cell(sheet, 'E15', дата)),
                    ('direct_cell_access', lambda: setattr(sheet['E15'], 'value', дата)),
                    ('cell_coordinate', lambda: setattr(sheet.cell(15, 5), 'value', дата))
                ]

                for method_name, method_func in methods:
                    try:
                        result = method_func()
                        actual_value = sheet['E15'].value

                        if actual_value and str(actual_value).strip() == str(дата).strip():
                            success_count += 1
                            break
                        else:
                            print(f"[DEBUG] ✗ Неуспіх методом {method_name}")

                    except Exception as e:
                        print(f"[DEBUG] Помилка методом {method_name}: {e}")
            else:
                print(f"[DEBUG] Дата пуста або None, пропускаємо")

            # Додаткова діагностика шаблону
            target_cells = ['D12', 'E14', 'E15']
            for cell_ref in target_cells:
                try:
                    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
                    col_letter, row = coordinate_from_string(cell_ref)
                    col_num = column_index_from_string(col_letter)

                    cell = sheet[cell_ref]
                    is_merged = is_merged_cell(sheet, row, col_num)

                    if is_merged:
                        # Знаходимо об'єднання
                        for merged_range in sheet.merged_cells.ranges:
                            if (merged_range.min_row <= row <= merged_range.max_row and
                                    merged_range.min_col <= col_num <= merged_range.max_col):
                                break

                except Exception as e:
                    print(f"[DEBUG] Помилка діагностики ячейки {cell_ref}: {e}")

            return success_count > 0

        except Exception as e:
            print(f"[ERROR] Критична помилка при заповненні основних даних: {e}")
            import traceback
            traceback.print_exc()
            return False

    try:
        template_path = "ШАБЛОН_кошторис_розумний.xlsx"
        output_path = f"кошторис_заповнений.xlsx"

        # Перевіряємо чи існує шаблон
        if not os.path.exists(template_path):
            print(f"[WARNING] Шаблон {template_path} не знайдено. Створюємо простий файл.")
            return save_koshtorys_to_text(захід, адреса, дата, товари, загальна_сума, сума_прописом)

        # Копіюємо шаблон
        shutil.copy2(template_path, output_path)

        if EXCEL_AVAILABLE:
            # Відкриваємо скопійований файл
            workbook = load_workbook(output_path)
            sheet = workbook.active

            # Вызовите эту функцию после открытия шаблона:
            debug_template_structure(sheet)

            # Заповнюємо основні дані відповідно до вказаних ячейок
            improved_fill_main_data(sheet, захід, адреса, дата)

            # Робимо жирний шрифт у злитій клітинці C13:E13
            bold_font = Font(bold=True)
            for col in ['C', 'D', 'E']:
                cell = sheet[f'{col}13']
                # Зберігаємо всі інші властивості шрифту, але робимо жирним
                cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=True,  # Примусово жирний
                    italic=cell.font.italic,
                    color=cell.font.color
                )

            print(f"[DEBUG] Зроблено жирний шрифт в ячейках C13:E13")

            # Вставляємо додаткові рядки для товарів (це автоматично зсуне весь контент)
            товари_count = len(товари)
            print(f"[DEBUG] Кількість товарів: {товари_count}")

            start_row = insert_rows_for_products(sheet, товари_count)

            # Заповнюємо товари
            base_b_number = 6  # Базове значення для стовпця B
            g_column_value = 2210  # Значення для стовпця G

            for i, товар_data in enumerate(товари):
                current_row = start_row + i

                кількість_num = convert_to_number(товар_data["кількість"])
                ціна_num = convert_to_number(товар_data["ціна за одиницю"])

                # НОВОЕ: Рассчитываем сумму по товару
                сума_товару = кількість_num * ціна_num

                # B - нумерація (6, 7, 8...)
                safe_set_cell_value(sheet, current_row, 2, base_b_number + i)  # B колонка = 2
                print(f"[DEBUG] B{current_row} = {base_b_number + i}")

                # C - назви товарів
                safe_set_cell_value(sheet, current_row, 3, товар_data["товар"])  # C колонка = 3

                # G - значення 2210 (жирним шрифтом)
                if safe_set_cell_value(sheet, current_row, 7, g_column_value):  # G колонка = 7
                    try:
                        g_cell = sheet.cell(current_row, 7)
                        if not is_merged_cell(sheet, current_row, 7):
                            if g_cell.font:
                                g_cell.font = Font(
                                    name=g_cell.font.name,
                                    size=g_cell.font.size,
                                    bold=True,
                                    italic=g_cell.font.italic,
                                    color=g_cell.font.color
                                )
                            else:
                                g_cell.font = Font(bold=True)
                    except:
                        pass
                print(f"[DEBUG] G{current_row} = {g_column_value} (жирним)")

                # H - кількість
                safe_set_cell_value(sheet, current_row, 8, кількість_num)  # H колонка = 8

                # J - ціна за одиницю
                safe_set_cell_value(sheet, current_row, 10, ціна_num)  # J колонка = 10

                # K - НОВОЕ: сума по товару (кількість × ціна)
                if safe_set_cell_value(sheet, current_row, 11, сума_товару):  # K колонка = 11
                    print(f"[DEBUG] K{current_row} = {сума_товару:.2f} (сума товару)")
                else:
                    print(f"[WARNING] Не вдалося записати суму товару в K{current_row}")

                # L - НОВОЕ: дублируем сумму товара в столбец L
                if safe_set_cell_value(sheet, current_row, 12, сума_товару):  # L колонка = 12
                    print(f"[DEBUG] L{current_row} = {сума_товару:.2f} (дубль суми товару)")
                else:
                    print(f"[WARNING] Не вдалося записати суму товару в L{current_row}")

                print(f"[DEBUG] Рядок {current_row}: товар='{товар_data['товар']}' (C{current_row}), "
                      f"кількість={кількість_num} (H{current_row}), ціна={ціна_num} (J{current_row}), "
                      f"сума={сума_товару:.2f} (K{current_row}, L{current_row})")

            # ИСПРАВЛЕНО: Ищем строку с "у т. ч. за КЕКВ:" и "2210"
            total_sum_row = None

            # Ищем строку где в столбцах C-F есть "у т. ч. за КЕКВ:" а в G есть "2210"
            for row in range(start_row + товари_count, start_row + товари_count + 20):
                try:
                    # Проверяем объединенные ячейки C-F на наличие "у т. ч. за КЕКВ:"
                    kekv_found = False
                    for col in ['C', 'D', 'E', 'F']:
                        cell_value = sheet[f'{col}{row}'].value
                        if cell_value and "у т. ч. за КЕКВ:" in str(cell_value):
                            kekv_found = True
                            break

                    # Проверяем столбец G на наличие "2210"
                    g_cell_value = sheet[f'G{row}'].value
                    if kekv_found and g_cell_value and "2210" in str(g_cell_value):
                        total_sum_row = row
                        print(f"[DEBUG] Найдена строка с КЕКВ:2210 в строке {row}")
                        break

                except Exception as e:
                    print(f"[DEBUG] Ошибка при поиске КЕКВ в строке {row}: {e}")
                    continue

            # Если нашли строку с КЕКВ:2210, записываем туда общую сумму
            if total_sum_row:
                if safe_set_cell_value(sheet, total_sum_row, 11, загальна_сума):  # K колонка = 11
                    print(f"[DEBUG] Загальна сума {загальна_сума:.2f} записана в K{total_sum_row} (строка с КЕКВ:2210)")
                else:
                    print(f"[WARNING] Не вдалося записати суму в K{total_sum_row}")

                # НОВОЕ: Дублируем общую сумму в столбец L
                if safe_set_cell_value(sheet, total_sum_row, 12, загальна_сума):  # L колонка = 12
                    print(f"[DEBUG] Загальна сума {загальна_сума:.2f} продубльована в L{total_sum_row}")
                else:
                    print(f"[WARNING] Не вдалося записати суму в L{total_sum_row}")
            else:
                print(f"[WARNING] Не знайдено строку з 'у т. ч. за КЕКВ:' та '2210'")

            # НОВОЕ: Продолжаем нумерацию в столбце B для перенесенных строк (33-36)
            # Нумерация продолжается после товаров
            next_b_number = base_b_number + товари_count

            # Заполняем нумерацию для строк которые были перенесены вниз (примерно 33-36)
            for offset in range(4):  # 4 строки (33, 34, 35, 36)
                current_row = start_row + товари_count + offset
                try:
                    # Проверяем есть ли контент в этой строке (не пустая ли)
                    has_content = False
                    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'J']:
                        if sheet[f'{col}{current_row}'].value:
                            has_content = True
                            break

                    if has_content:
                        if safe_set_cell_value(sheet, current_row, 2, next_b_number):  # B колонка = 2
                            print(f"[DEBUG] B{current_row} = {next_b_number} (продолжение нумерации)")
                            next_b_number += 1

                except Exception as e:
                    print(f"[DEBUG] Ошибка при нумерации строки {current_row}: {e}")
                    continue

            # Зберігаємо файл
            workbook.save(output_path)
            workbook.close()

            print(f"[SUCCESS] Кошторис збережено у файл: {output_path}")
            messagebox.showinfo("Успіх", f"Кошторис успішно збережено у файл:\n{output_path}\n\n"
                                         f"Додано {товари_count} товарів з правильним форматуванням!\n"
                                         f"Суми по товарах додані в стовпці K та L (починаючи з K32, L32)\n"
                                         f"Загальна сума записана в строку з КЕКВ:2210 (стовпці K та L)\n"
                                         f"Нумерація продовжена для всіх рядків")
            return output_path
        else:
            # Якщо openpyxl недоступний, створюємо текстовий файл
            return save_koshtorys_to_text(захід, адреса, дата, товари, загальна_сума, сума_прописом)

    except Exception as e:
        error_msg = f"Помилка при збереженні Excel файлу: {str(e)}"
        print(f"[ERROR] {error_msg}")
        traceback.print_exc()
        messagebox.showerror("Помилка", error_msg)
        # Якщо не вдалося зберегти Excel, спробуємо текстовий формат
        # return save_koshtorys_to_text(захід, адреса, дата, товари, загальна_сума, сума_прописом)



def save_koshtorys_to_text(захід, адреса, дата, товари, загальна_сума, сума_прописом):
    """Зберігає кошторис у текстовий файл (резервний варіант)"""
    return None


def fill_koshtorys(document_blocks):
    """Основна функція заповнення кошторису"""
    try:
        if not document_blocks:
            messagebox.showwarning("Увага", "Не знайдено даних для заповнення кошторису.")
            return False

        # Отримуємо дані з першого блоку для загальних полів
        first_block = document_blocks[0]
        first_entries = first_block.get("entries", {})

        # Отримуємо основні поля
        захід = get_entry_value(first_entries, "захід")
        адреса = get_entry_value(first_entries, "адреса")
        дата = get_entry_value(first_entries, "дата")

        # Отримуємо список товарів
        товари = get_товари_from_blocks(document_blocks)

        if not товари:
            messagebox.showwarning("Увага", "Не знайдено товарів для заповнення кошторису.\n"
                                            "Переконайтесь, що у блоках договорів заповнені поля:\n"
                                            "- товар/назва/найменування/предмет\n"
                                            "- кількість\n"
                                            "- ціна за одиницю")
            return False

        print(f"[DEBUG] Знайдено товарів: {len(товари)}")
        print(f"[DEBUG] Захід: {захід}")
        print(f"[DEBUG] Адреса: {адреса}")
        print(f"[DEBUG] Дата: {дата}")

        # Виводимо інформацію про товари та рахуємо загальну суму
        загальна_сума = 0.0
        товари_info = []

        for i, товар_data in enumerate(товари):
            кількість_num = convert_to_number(товар_data["кількість"])
            ціна_num = convert_to_number(товар_data["ціна за одиницю"])
            сума_товару = кількість_num * ціна_num
            загальна_сума += сума_товару

            товар_info = (f"{i + 1}. {товар_data['товар']} - "
                          f"{кількість_num} {товар_data['одиниця виміру']} × "
                          f"{ціна_num:.2f} грн = {сума_товару:.2f} грн")
            товари_info.append(товар_info)

            print(f"[DEBUG] Товар {i + 1}: {товар_data['товар']}, "
                  f"кількість: {кількість_num} {товар_data['одиниця виміру']}, "
                  f"ціна: {ціна_num:.2f}, разом: {сума_товару:.2f}")

        print(f"[DEBUG] Загальна сума: {загальна_сума:.2f}")

        # Формуємо детальне повідомлення
        товари_список = "\n".join(товари_info)
        сума_прописом = number_to_ukrainian_text(загальна_сума).capitalize()

        detail_message = (f"ДАНІ ДЛЯ КОШТОРИСУ:\n\n"
                          f"Захід: {захід}\n"
                          f"Місце проведення: {адреса}\n"
                          f"Дата: {дата}\n\n"
                          f"ТОВАРИ ТА ПОСЛУГИ:\n{товари_список}\n\n"
                          f"ЗАГАЛЬНА СУМА: {загальна_сума:.2f} грн\n"
                          f"Сума прописом: {сума_прописом}")

        # Зберігаємо кошторис у файл з правильним форматуванням
        saved_file = save_koshtorys_to_excel(захід, адреса, дата, товари, загальна_сума, сума_прописом)

        return saved_file is not None

    except Exception as e:
        error_msg = f"Помилка при обробці кошторису: {str(e)}"
        print(f"[ERROR] {error_msg}")
        traceback.print_exc()
        messagebox.showerror("Помилка", error_msg)
        return False


def generate_koshtorys_only(document_blocks):
    """Функція для генерації тільки кошторису (для окремої кнопки)"""
    return fill_koshtorys(document_blocks)