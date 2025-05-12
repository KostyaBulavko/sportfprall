import os
import re
import tkinter.filedialog as filedialog
import tkinter.messagebox as messagebox
import traceback
import datetime
import sys
from openpyxl import load_workbook

# Константы для ячеек в кошторис.xlsx - теперь в виде шаблонов для нескольких договоров
# Для первого договора ячейки остаются теми же
KOSHTORYS_CELLS_TEMPLATE = {
    "назва_заходу": "D12",       
    "адреса": "E14",             
    "дата": "E15",               
    "товар": "C{row}",             
    "кількість": "H{row}",
    "ціна за одиницю": "J{row}",
    "разом": "K{row}",
    "всього_сума_словами": "B45",
    "сума_словами": "H6",
    "загальна_сума": "K41"  # Добавлена ячейка с общей суммой
}

# Определяем поля, которые должны быть числами в Excel
NUMERIC_FIELDS = ["кількість", "ціна за одиницю", "разом"]

# Начальная строка для первого договора
BASE_ROW = 32

# Добавляем константу для пути к файлу кошторис.xlsx
KOSHTORYS_PATH = "ШАБЛОН_кошторис.xlsx"

# Функция для записи ошибок в лог
def log_error(exc_type, exc_value, exc_traceback, error_log="error.txt"):
    try:
        with open(error_log, "a", encoding="utf-8") as f:
            f.write(f"\n[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]\n")
            traceback.print_exception(exc_type, exc_value, exc_traceback, file=f)
        # Показать сообщение об ошибке
        messagebox.showerror("Ошибка", f"Произошла ошибка: {str(exc_value)}\nПодробности в файле {error_log}")
    except Exception as ex:
        print(f"Ошибка при логировании: {str(ex)}")

# Функция для преобразования строки в число для Excel
def convert_to_excel_number(value_str):
    # Удаляем пробелы и заменяем запятые на точки
    clean_str = value_str.replace(" ", "").replace(",", ".")
    
    # Извлекаем только цифры и одну десятичную точку
    matches = re.search(r'(\d+\.?\d*)', clean_str)
    if matches:
        number_str = matches.group(1)
        try:
            # Преобразуем в число
            return float(number_str)
        except ValueError:
            pass
    
    # Если не удалось преобразовать, возвращаем исходную строку
    return value_str

# Функция для конвертации числа в текст прописью (украинский вариант)
def number_to_text(number, currency="гривень"):
    # Это упрощенная версия функции - на практике должна быть полная реализация 
    # преобразования числа в текст на украинском языке
    
    # В реальном приложении здесь будет полноценная функция для преобразования числа в текст
    # Например: 37378.00 -> "Тридцять сім тисяч триста сімдесят вісім гривень, 00 коп"
    
    # Заглушка для примера
    hrn = int(number)  # Целая часть в гривнах
    kop = int((number - hrn) * 100)  # Копейки
    
    # В реальном коде здесь будет сложная логика преобразования чисел в слова
    # на украинском языке с учетом склонений
    
    # Пример вывода для демонстрации формата
    return f"Тридцять сім тисяч триста сімдесят вісім {currency}, {kop:02d} коп"

# Функция для получения текущих ячейки для конкретной строки
def get_cells_for_row(row):
    cells = {}
    for key, template in KOSHTORYS_CELLS_TEMPLATE.items():
        if "{row}" in template:
            cells[key] = template.format(row=row)
        else:
            cells[key] = template
    return cells

# Функция для заполнения файла кошторис.xlsx
def fill_koshtorys(document_blocks):
    try:
        # Проверяем, есть ли документы
        if not document_blocks:
            messagebox.showwarning("Увага", "Не додано жодного кошториса")
            return False
            
        # Проверяем существование файла
        if not os.path.exists(KOSHTORYS_PATH):
            # Спрашиваем пользователя о местоположении файла
            koshtorys_file = filedialog.askopenfilename(
                title="Выберите файл кошторис.xlsx", 
                filetypes=[("Excel Files", "*.xlsx")]
            )
            if not koshtorys_file:
                messagebox.showwarning("Увага", "Файл кошторис.xlsx не вибран.")
                return False
        else:
            koshtorys_file = KOSHTORYS_PATH
            
        try:
            # Загружаем существующий файл
            wb = load_workbook(koshtorys_file)
            ws = wb.active
            
            # Функция для определения, есть ли ПДВ в любом из договоров
            def check_pdv_status(blocks):
                for block in blocks:
                    entries = block["entries"]
                    if "пдв" in entries and entries["пдв"].get():
                        pdv_text = entries["пдв"].get().strip().lower()
                        if "пдв" in pdv_text and "без" not in pdv_text:
                            return "з ПДВ"
                return "без ПДВ"
            
            # Получаем статус ПДВ
            pdv_status = check_pdv_status(document_blocks)
            
            # Функция для безопасной записи в ячейку, учитывая возможные объединенные ячейки
            def safe_write_cell(cell_address, value, is_numeric=False):
                try:
                    # Преобразуем в число, если это числовое поле
                    final_value = convert_to_excel_number(value) if is_numeric else value
                    
                    # Записываем значение
                    ws[cell_address] = final_value
                except AttributeError:
                    # Если ячейка объединенная, найдем основную ячейку из объединенных ячеек
                    for merged_range in ws.merged_cells.ranges:
                        if cell_address in merged_range:
                            # Используем верхнюю левую ячейку объединенного диапазона
                            ws[merged_range.start_cell.coordinate] = final_value if is_numeric else value
                            return
                    # Если ячейка не найдена в объединенных диапазонах, выводим предупреждение
                    messagebox.showwarning("Увага", f"Не вдалось записати значення в ячейку {cell_address}")
            
            # Функция для безопасного чтения из ячейки
            def safe_read_cell(cell_address):
                try:
                    return ws[cell_address].value
                except:
                    # Если ячейка объединенная, найдем основную ячейку из объединенных ячеек
                    for merged_range in ws.merged_cells.ranges:
                        if cell_address in merged_range:
                            # Используем верхнюю левую ячейку объединенного диапазона
                            return ws[merged_range.start_cell.coordinate].value
                    return None
            
            # Заполняем общие данные (берем из первого блока)
            if document_blocks:
                first_entry = document_blocks[0]["entries"]
                common_cells = get_cells_for_row(BASE_ROW)  # Получаем ячейки с общими данными
                safe_write_cell(common_cells["назва_заходу"], first_entry["захід"].get())           
                safe_write_cell(common_cells["адреса"], first_entry["адреса"].get())
                safe_write_cell(common_cells["дата"], first_entry["дата"].get())
                
                # Заполняем данные по каждому договору
                for i, block in enumerate(document_blocks):
                    entries = block["entries"]
                    # Вычисляем строку для текущего договора (первый договор - BASE_ROW, второй - BASE_ROW+1 и т.д.)
                    current_row = BASE_ROW + i
                    cells = get_cells_for_row(current_row)
                    
                    # Заполняем данные для текущего договора
                    safe_write_cell(cells["товар"], entries["товар"].get())
                    
                    # Числовые поля записываем как числа
                    for field in NUMERIC_FIELDS:
                        if field in cells and field in entries:
                            safe_write_cell(cells[field], entries[field].get(), is_numeric=True)
                
                # Вместо считывания K41, суммируем значения из ячеек K27:K40
                total_sum = 0.0
                
                # Суммируем значения в ячейках от K27 до K40
                for row in range(27, 41):
                    cell_value = safe_read_cell(f"K{row}")
                    if cell_value is not None:
                        try:
                            # Попытка преобразовать в число
                            if isinstance(cell_value, (int, float)):
                                total_sum += cell_value
                            else:
                                # Если значение не числовое, преобразуем строку в число
                                number_value = convert_to_excel_number(str(cell_value))
                                if isinstance(number_value, (int, float)):
                                    total_sum += number_value
                        except:
                            # В случае ошибки преобразования, игнорируем это значение
                            pass
                
                # Записываем вычисленную общую сумму в ячейку K41
                safe_write_cell("K41", str(total_sum), is_numeric=True)

                #------функцію конвертації числа в текст українською мовою------------
                def number_to_ukrainian_text(amount):
                    UNITS = (
                        ("нуль",),
                        ("одна", "один"),
                        ("дві", "два"),
                        ("три", "три"),
                        ("чотири", "чотири"),
                        ("п’ять",),
                        ("шість",),
                        ("сім",),
                        ("вісім",),
                        ("дев’ять",),
                    )
                    
                    TENS = (
                        "", "десять", "двадцять", "тридцять", "сорок", "п’ятдесят",
                        "шістдесят", "сімдесят", "вісімдесят", "дев’яносто"
                    )
                    
                    TEENS = (
                        "десять", "одинадцять", "дванадцять", "тринадцять", "чотирнадцять",
                        "п’ятнадцять", "шістнадцять", "сімнадцять", "вісімнадцять", "дев’ятнадцять"
                    )
                    
                    HUNDREDS = (
                        "", "сто", "двісті", "триста", "чотириста",
                        "п’ятсот", "шістсот", "сімсот", "вісімсот", "дев’ятсот"
                    )
                    
                    ORDERS = [
                        (("гривня", "гривні", "гривень"), "f"),
                        (("тисяча", "тисячі", "тисяч"), "f"),
                        (("мільйон", "мільйони", "мільйонів"), "m"),
                        (("мільярд", "мільярди", "мільярдів"), "m")
                    ]

                    def get_plural_form(n):
                        if n % 10 == 1 and n % 100 != 11:
                            return 0
                        elif 2 <= n % 10 <= 4 and not (12 <= n % 100 <= 14):
                            return 1
                        return 2

                    def convert_triplet(n, gender):
                        result = []
                        n = int(n)
                        if n == 0:
                            return []

                        hundreds = n // 100
                        tens_units = n % 100
                        tens = tens_units // 10
                        units = tens_units % 10

                        result.append(HUNDREDS[hundreds])

                        if tens_units >= 10 and tens_units < 20:
                            result.append(TEENS[tens_units - 10])
                        else:
                            result.append(TENS[tens])
                            if units > 0:
                                if units < 3:
                                    result.append(UNITS[units][0 if gender == "f" else 1])
                                else:
                                    result.append(UNITS[units][0])

                        return [word for word in result if word]

                    integer_part = int(amount)
                    fractional_part = round((amount - integer_part) * 100)

                    if integer_part == 0:
                        words = ["нуль"]
                    else:
                        words = []
                        order = 0
                        while integer_part > 0 and order < len(ORDERS):
                            n = integer_part % 1000
                            if n > 0:
                                group_words = convert_triplet(n, ORDERS[order][1])
                                group_name = ORDERS[order][0][get_plural_form(n)]
                                words = group_words + [group_name] + words
                            integer_part //= 1000
                            order += 1

                    kop = f"{fractional_part:02d}"
                    kop_words = f"{kop} коп."

                    return f"{' '.join(words)}, {kop_words}"



                # Если значение получено, используем его для формирования итоговых сумм
                if total_sum > 0:
                    # Получаем сумму прописью
                    suma_propysom = number_to_ukrainian_text(total_sum)

                    # Робимо перше слово після дужки з великої літери
                    suma_propysom = suma_propysom[0].upper() + suma_propysom[1:]
                    
                    # Формируем строку для B45 - сумма прописью со скобками и статусом ПДВ
                    vsogo_text = f"Всього: ({suma_propysom}, {pdv_status})"
                    safe_write_cell("B45", vsogo_text)

                    # Формируем строку для H6 - только сумма прописью
                    vsogo_text_H6 = f"({suma_propysom}, {pdv_status})"
                    safe_write_cell("H6", vsogo_text_H6)

                    # Формируем строку для H6 - только сумма прописью
                    #safe_write_cell("H6", suma_propysom)

                else:
                    # Если не удалось получить значение, выводим предупреждение
                    messagebox.showwarning("Увага", "Не вдалось розрахувати загальну суму з ячеек K27:K40")
                    
                    # В этом случае используем данные из полей формы, как было раньше
                    suma_propysom = ""
                    
                    # Проверяем, есть ли поле "сума прописом" в записях
                    if "сума прописом" in first_entry:
                        suma_propysom = first_entry["сума прописом"].get()
                    
                    # Если сумма прописью не найдена, используем поле "сума"
                    if not suma_propysom and "сума" in first_entry:
                        suma_propysom = first_entry["сума"].get()
                    
                    # Заполняем ячейку H6 суммой прописью
                    safe_write_cell ("H6", suma_propysom)

                    # Заполняем ячейку B45 с текстом "Всього: " + сумма прописью + ПДВ статус
                    vsogo_text = f"Всього: ({suma_propysom}, {pdv_status})"
                    safe_write_cell("B45", vsogo_text)


            # Сохраняем изменения
            #save_path = os.path.join(os.path.dirname(koshtorys_file), f"заповнений_кошторис_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            save_path = os.path.join(os.path.dirname(koshtorys_file), f"Кошторис.xlsx")

            wb.save(save_path)
            
            messagebox.showinfo("Успех", f"Кошторис успешно заполнен и сохранен как:\n{save_path}")
            return True
            
        except Exception as e:
            log_error(type(e), e, sys.exc_info()[2])
            messagebox.showerror("Ошибка", f"Ошибка при заполнении кошториса: {str(e)}")
            return False
            
    except Exception as e:
        log_error(type(e), e, sys.exc_info()[2])
        messagebox.showerror("Ошибка", f"Неизвестная ошибка: {str(e)}")
        return False

def set_koshtorys_path(path):
    """Установить новый путь к файлу кошториса"""
    global KOSHTORYS_PATH
    KOSHTORYS_PATH = path
    return KOSHTORYS_PATH

def get_koshtorys_path():
    """Получить текущий путь к файлу кошториса"""
    return KOSHTORYS_PATH

def set_base_row(row):
    """Установить базовую строку для начала заполнения договоров"""
    global BASE_ROW
    BASE_ROW = row
    return BASE_ROW

def get_base_row():
    """Получить текущую базовую строку"""
    return BASE_ROW

# Функция для обновления шаблонов ячеек кошториса
def update_koshtorys_cells_template(cells_dict):
    """Обновить шаблоны соответствия ячеек в файле кошториса"""
    global KOSHTORYS_CELLS_TEMPLATE
    KOSHTORYS_CELLS_TEMPLATE.update(cells_dict)
    return KOSHTORYS_CELLS_TEMPLATE

# Функция для обновления списка числовых полей
def update_numeric_fields(fields_list):
    """Обновить список полей, которые должны быть числами в Excel"""
    global NUMERIC_FIELDS
    NUMERIC_FIELDS = fields_list
    return NUMERIC_FIELDS