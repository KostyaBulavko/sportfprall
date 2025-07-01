# excel_data_processor.py
"""
Обработка данных и работа со строками в Excel
"""

import os
import re
from openpyxl import Workbook, load_workbook
from excel_config import get_ordered_headers, get_headers_config
from excel_formatters import format_headers, apply_column_widths


def ensure_file_structure(filename, fields_list):
    """
    Проверяет существование файла и создает его со структурой заголовков если нужно
    Возвращает рабочий лист
    """
    headers = get_ordered_headers(fields_list)

    if os.path.exists(filename):
        try:
            wb = load_workbook(filename)
            ws = wb.active

            # Проверяем, есть ли заголовки
            if ws.max_row == 0 or not ws[1][0].value:
                # Файл пустой, добавляем заголовки
                ws.append(headers)
                format_headers(ws, headers)
                apply_column_widths(ws, fields_list)
                wb.save(filename)
                print(f"[INFO] Добавлены заголовки в существующий файл: {filename}")
            else:
                # Проверяем соответствие заголовков
                existing_headers = [cell.value for cell in ws[1] if cell.value]
                if existing_headers != headers:
                    print(f"[WARNING] Заголовки в файле не соответствуют ожидаемым")
                    print(f"[WARNING] Ожидаемые: {headers}")
                    print(f"[WARNING] Существующие: {existing_headers}")
                    # Можно добавить логику обновления заголовков здесь

            return wb, ws

        except Exception as e:
            print(f"[ERROR] Ошибка при открытии файла {filename}: {e}")
            # Создаем новый файл
            return create_new_file(filename, fields_list)
    else:
        # Файл не существует, создаем новый
        return create_new_file(filename, fields_list)


def create_new_file(filename, fields_list):
    """Создает новый файл с заголовками"""
    headers = get_ordered_headers(fields_list)
    wb = Workbook()
    ws = wb.active
    ws.title = "База даних"

    # Добавляем заголовки
    ws.append(headers)
    format_headers(ws, headers)
    apply_column_widths(ws, fields_list)

    wb.save(filename)
    print(f"[INFO] Создан новый файл: {filename}")
    return wb, ws


def find_next_row(ws):
    """Находит следующую пустую строку для вставки данных"""
    return ws.max_row + 1


def find_event_product_row(ws, event_number, product_name, headers):
    """
    Ищет строку с указанным номером события и товаром
    Возвращает номер строки или None если не найдено
    """
    if not event_number or event_number == "":
        return None

    # Проверяем, что есть данные
    if ws.max_row < 2:  # Нет данных, только заголовки или пустой лист
        return None

    # Находим индексы нужных колонок
    event_number_col = None
    product_col = None

    for col_idx, header in enumerate(headers, 1):
        if header.lower() == "номер заходу":
            event_number_col = col_idx
        elif header.lower() == "товар":
            product_col = col_idx

    if event_number_col is None:
        print(f"[ERROR] Не найдена колонка 'Номер заходу'")
        return None

    # Ищем строку с совпадающим номером события и товаром
    for row in range(2, ws.max_row + 1):  # Начинаем с 2-й строки (после заголовков)
        event_cell_value = ws.cell(row=row, column=event_number_col).value

        # Проверяем номер события
        if str(event_cell_value) == str(event_number):
            # Если есть колонка товара, проверяем и её
            if product_col is not None and product_name:
                product_cell_value = ws.cell(row=row, column=product_col).value
                if str(product_cell_value or "").lower() == str(product_name or "").lower():
                    print(
                        f"[DEBUG] Найдена существующая запись для события {event_number}, товар '{product_name}' в строке {row}")
                    return row
            else:
                # Если нет колонки товара или товар не указан, берем первое совпадение по событию
                print(f"[DEBUG] Найдена существующая запись для события {event_number} в строке {row}")
                return row

    return None


def find_all_event_rows(ws, event_number, headers):
    """
    Находит все строки с указанным номером события
    Возвращает список номеров строк
    """
    if not event_number or event_number == "":
        return []

    # Проверяем, что есть данные
    if ws.max_row < 2:
        return []

    # Находим индекс колонки с номером события
    event_number_col = None
    for col_idx, header in enumerate(headers, 1):
        if header.lower() == "номер заходу":
            event_number_col = col_idx
            break

    if event_number_col is None:
        return []

    # Ищем все строки с этим номером события
    matching_rows = []
    for row in range(2, ws.max_row + 1):
        event_cell_value = ws.cell(row=row, column=event_number_col).value
        if str(event_cell_value) == str(event_number):
            matching_rows.append(row)

    return matching_rows


def update_existing_row(ws, row_number, row_data, headers):
    """
    Обновляет существующую строку данными
    """
    print(f"[DEBUG] Обновляем строку {row_number}")

    for col_num, value in enumerate(row_data, 1):
        if col_num <= len(headers):  # Проверяем, что не выходим за границы заголовков
            old_value = ws.cell(row=row_number, column=col_num).value
            ws.cell(row=row_number, column=col_num, value=value)

            if str(old_value) != str(value):
                header_name = headers[col_num - 1] if col_num <= len(headers) else f"Колонка_{col_num}"
                print(f"[DEBUG] Обновлено '{header_name}': '{old_value}' -> '{value}'")


def add_new_row(ws, row_data):
    """
    Добавляет новую строку в конец таблицы
    """
    new_row = ws.max_row + 1
    print(f"[DEBUG] Добавляем новую строку {new_row}")

    for col_num, value in enumerate(row_data, 1):
        ws.cell(row=new_row, column=col_num, value=value)

    return new_row


def process_data_row(ws, row_data, headers, event_number, product_name=None):
    """
    Обрабатывает строку данных с учетом логики "событие + товар"
    - Если найдена строка с таким же событием и товаром - обновляет её
    - Если событие есть, но товар другой - добавляет новую строку
    - Если события нет - добавляет новую строку

    Возвращает (действие, номер_строки) где действие = 'updated' или 'added'
    """
    if event_number and event_number != "":
        # Ищем существующую запись с таким же событием и товаром
        existing_row = find_event_product_row(ws, event_number, product_name, headers)

        if existing_row:
            # Обновляем существующую запись
            update_existing_row(ws, existing_row, row_data, headers)
            return 'updated', existing_row
        else:
            # Событие есть, но товар другой или не найден - добавляем новую строку
            new_row = add_new_row(ws, row_data)

            # Получаем все строки с этим событием для информации
            all_event_rows = find_all_event_rows(ws, event_number, headers)
            if all_event_rows:
                print(f"[INFO] Для события {event_number} уже есть {len(all_event_rows)} записей, добавляем новую")

            return 'added', new_row
    else:
        # Нет номера события - добавляем как новую запись
        print(f"[WARNING] Нет номера события, добавляем как новую запись")
        new_row = add_new_row(ws, row_data)
        return 'added', new_row


def get_product_name_from_row_data(row_data, headers):
    """
    Извлекает название товара из данных строки
    """
    try:
        # Ищем индекс колонки "товар"
        product_index = None
        for idx, header in enumerate(headers):
            if header.lower() == "товар":
                product_index = idx
                break

        if product_index is not None and product_index < len(row_data):
            return row_data[product_index]
        else:
            return None
    except Exception as e:
        print(f"[ERROR] Ошибка при извлечении названия товара: {e}")
        return None