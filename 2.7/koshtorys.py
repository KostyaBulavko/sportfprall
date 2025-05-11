import os
import re
import tkinter.filedialog as filedialog
import tkinter.messagebox as messagebox
import traceback
import datetime
import sys
from openpyxl import load_workbook

# Константы для ячеек в кошторис.xlsx - теперь в виде шаблонов для нескольких договоров
# Для первого договора ячейки остаются теми же
KOSHTORYS_CELLS_TEMPLATE = {
    "назва_заходу": "D12",       
    "адреса": "E14",             
    "дата": "E15",               
    "товар": "C{row}",             
    "кількість": "H{row}",
    "ціна за одиницю": "J{row}",
    "разом": "K{row}",
}

# Определяем поля, которые должны быть числами в Excel
NUMERIC_FIELDS = ["кількість", "ціна за одиницю", "разом"]

# Начальная строка для первого договора
BASE_ROW = 30

# Добавляем константу для пути к файлу кошторис.xlsx
KOSHTORYS_PATH = "кошторис.xlsx"

# Функция для записи ошибок в лог
def log_error(exc_type, exc_value, exc_traceback, error_log="error.txt"):
    try:
        with open(error_log, "a", encoding="utf-8") as f:
            f.write(f"\n[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]\n")
            traceback.print_exception(exc_type, exc_value, exc_traceback, file=f)
        # Показать сообщение об ошибке
        messagebox.showerror("Ошибка", f"Произошла ошибка: {str(exc_value)}\nПодробности в файле {error_log}")
    except Exception as ex:
        print(f"Ошибка при логировании: {str(ex)}")

# Функция для преобразования строки в число для Excel
def convert_to_excel_number(value_str):
    # Удаляем пробелы и заменяем запятые на точки
    clean_str = value_str.replace(" ", "").replace(",", ".")
    
    # Извлекаем только цифры и одну десятичную точку
    matches = re.search(r'(\d+\.?\d*)', clean_str)
    if matches:
        number_str = matches.group(1)
        try:
            # Преобразуем в число
            return float(number_str)
        except ValueError:
            pass
    
    # Если не удалось преобразовать, возвращаем исходную строку
    return value_str

# Функция для получения текущих ячейки для конкретной строки
def get_cells_for_row(row):
    cells = {}
    for key, template in KOSHTORYS_CELLS_TEMPLATE.items():
        if "{row}" in template:
            cells[key] = template.format(row=row)
        else:
            cells[key] = template
    return cells

# Функция для заполнения файла кошторис.xlsx
def fill_koshtorys(document_blocks):
    try:
        # Проверяем, есть ли документы
        if not document_blocks:
            messagebox.showwarning("Внимание", "Не добавлено ни одного документа.")
            return False
            
        # Проверяем существование файла
        if not os.path.exists(KOSHTORYS_PATH):
            # Спрашиваем пользователя о местоположении файла
            koshtorys_file = filedialog.askopenfilename(
                title="Выберите файл кошторис.xlsx", 
                filetypes=[("Excel Files", "*.xlsx")]
            )
            if not koshtorys_file:
                messagebox.showwarning("Внимание", "Файл кошторис.xlsx не выбран.")
                return False
        else:
            koshtorys_file = KOSHTORYS_PATH
            
        try:
            # Загружаем существующий файл
            wb = load_workbook(koshtorys_file)
            ws = wb.active
            
            # Функция для безопасной записи в ячейку, учитывая возможные объединенные ячейки
            def safe_write_cell(cell_address, value, is_numeric=False):
                try:
                    # Преобразуем в число, если это числовое поле
                    final_value = convert_to_excel_number(value) if is_numeric else value
                    
                    # Записываем значение
                    ws[cell_address] = final_value
                except AttributeError:
                    # Если ячейка объединенная, найдем основную ячейку из объединенных ячеек
                    for merged_range in ws.merged_cells.ranges:
                        if cell_address in merged_range:
                            # Используем верхнюю левую ячейку объединенного диапазона
                            ws[merged_range.start_cell.coordinate] = final_value if is_numeric else value
                            return
                    # Если ячейка не найдена в объединенных диапазонах, выводим предупреждение
                    messagebox.showwarning("Внимание", f"Не удалось записать значение в ячейку {cell_address}")
            
            # Заполняем общие данные (берем из первого блока)
            if document_blocks:
                first_entry = document_blocks[0]["entries"]
                common_cells = get_cells_for_row(BASE_ROW)  # Получаем ячейки с общими данными
                safe_write_cell(common_cells["назва_заходу"], first_entry["захід"].get())           
                safe_write_cell(common_cells["адреса"], first_entry["адреса"].get())
                safe_write_cell(common_cells["дата"], first_entry["дата"].get())
                
                # Записываем сумму (текстовое поле) из первого договора
                common_cells = get_cells_for_row(BASE_ROW)  # Получаем ячейки
                if "сума" in common_cells:
                    safe_write_cell(common_cells["сума"], first_entry["сума"].get())
            
            # Заполняем данные по каждому договору
            for i, block in enumerate(document_blocks):
                entries = block["entries"]
                # Вычисляем строку для текущего договора (первый договор - BASE_ROW, второй - BASE_ROW+1 и т.д.)
                current_row = BASE_ROW + i
                cells = get_cells_for_row(current_row)
                
                # Заполняем данные для текущего договора
                safe_write_cell(cells["товар"], entries["товар"].get())
                
                # Числовые поля записываем как числа
                for field in NUMERIC_FIELDS:
                    if field in cells and field in entries:
                        safe_write_cell(cells[field], entries[field].get(), is_numeric=True)

            # Сохраняем изменения
            save_path = os.path.join(os.path.dirname(koshtorys_file), f"заповнений_кошторис_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb.save(save_path)
            
            messagebox.showinfo("Успех", f"Кошторис успешно заполнен и сохранен как:\n{save_path}")
            return True
            
        except Exception as e:
            log_error(type(e), e, sys.exc_info()[2])
            messagebox.showerror("Ошибка", f"Ошибка при заполнении кошториса: {str(e)}")
            return False
            
    except Exception as e:
        log_error(type(e), e, sys.exc_info()[2])
        messagebox.showerror("Ошибка", f"Неизвестная ошибка: {str(e)}")
        return False

def set_koshtorys_path(path):
    """Установить новый путь к файлу кошториса"""
    global KOSHTORYS_PATH
    KOSHTORYS_PATH = path
    return KOSHTORYS_PATH

def get_koshtorys_path():
    """Получить текущий путь к файлу кошториса"""
    return KOSHTORYS_PATH

def set_base_row(row):
    """Установить базовую строку для начала заполнения договоров"""
    global BASE_ROW
    BASE_ROW = row
    return BASE_ROW

def get_base_row():
    """Получить текущую базовую строку"""
    return BASE_ROW

# Функция для обновления шаблонов ячеек кошториса
def update_koshtorys_cells_template(cells_dict):
    """Обновить шаблоны соответствия ячеек в файле кошториса"""
    global KOSHTORYS_CELLS_TEMPLATE
    KOSHTORYS_CELLS_TEMPLATE.update(cells_dict)
    return KOSHTORYS_CELLS_TEMPLATE

# Функция для обновления списка числовых полей
def update_numeric_fields(fields_list):
    """Обновить список полей, которые должны быть числами в Excel"""
    global NUMERIC_FIELDS
    NUMERIC_FIELDS = fields_list
    return NUMERIC_FIELDS